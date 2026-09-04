# -*- coding: utf-8 -*-
"""Источники сторожевого дашборда: что тянем, откуда, как читаем.

Каждый источник забирается с браузерным User-Agent (без него climatereanalyzer
отвечает «Houston, we've had a problem»), кладётся в data/raw/<штамп>/ как есть,
и только потом разбирается. Сырое хранится дословно и с датой — если разбор
сломается через месяц, будет с чем сравнить.

При недоступности источника берётся последняя удачная копия, и это помечается:
дашборд обязан сказать, что данные несвежие, а не молча показать старое.
"""
import io
import json
import os
import re
import shutil
import time
import urllib.request
from datetime import datetime, date
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[2] / "data" / "enso"   # данные дашборда живут в data/enso/, код в tools/enso/
RAW = ROOT / "raw"
LAST = ROOT / "last_good"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"


def kuwait_url():
    """ERA5 в точке Кувейта (29.37°N, 47.98°E) через Open-Meteo, с начала года по вчера: адрес
    зависит от даты, поэтому это функция, а не строка. Климатологию 1991–2020 для той же
    точки строит gulf.py один раз и держит в кэше."""
    from datetime import timedelta
    t = date.today() - timedelta(days=1)
    return ("https://archive-api.open-meteo.com/v1/archive?latitude=29.37&longitude=47.98"
            f"&start_date={t.year}-01-01&end_date={t.isoformat()}"
            "&daily=temperature_2m_max,temperature_2m_min,temperature_2m_mean,precipitation_sum&timezone=UTC")


# name -> (url, kind); url может быть функцией — для адресов с датой
SOURCES = {
    "t2_world":   ("https://climatereanalyzer.org/clim/t2_daily/json/era5_world_t2_day.json", "cr_json"),
    "t2_nh":      ("https://climatereanalyzer.org/clim/t2_daily/json/era5_nh_t2_day.json", "cr_json"),
    "t2_sh":      ("https://climatereanalyzer.org/clim/t2_daily/json/era5_sh_t2_day.json", "cr_json"),
    "sst_world":  ("https://climatereanalyzer.org/clim/sst_daily/json_2clim/oisst2.1_world2_sst_day.json", "cr_json"),
    "sst_nino34": ("https://climatereanalyzer.org/clim/sst_daily/json_2clim/oisst2.1_nino3.4_sst_day.json", "cr_json"),
    "noaa_weekly": ("https://www.cpc.ncep.noaa.gov/data/indices/wksst9120.for", "noaa_weekly"),
    "oni":        ("https://www.cpc.ncep.noaa.gov/data/indices/oni.ascii.txt", "oni_txt"),
    "psl_nino34_monthly": ("https://psl.noaa.gov/data/correlation/nina34.anom.data", "psl_monthly"),
    # Продовольствие: единственный живой ряд без регистрации (ИСТОЧНИКИ.md, §4). CSV с шапкой
    # в четыре строки: Date, Food Price Index, Meat, Dairy, Cereals, Oils, Sugar; 2014-16 = 100.
    "fao_fpi": ("https://www.fao.org/media/docs/worldfoodsituationlibraries/default-document-library/"
                "food_price_indices_data.csv?download=true", "fao_csv"),

    # ── АТМОСФЕРА. До 4 сентября дашборд мерил только океан, а Эль-Ниньо — это связка океана
    # и воздуха: без атмосферы нельзя сказать, сцеплено событие или вода греется в одиночку,
    # и именно сцепка включает телесвязи (засухи, дожди, урожай). Владелец 04.09: «бери всё».
    "soi":        ("https://www.cpc.ncep.noaa.gov/data/indices/soi", "cpc_table"),
    "olr":        ("https://www.cpc.ncep.noaa.gov/data/indices/olr", "cpc_table"),
    "u850_west":  ("https://www.cpc.ncep.noaa.gov/data/indices/wpac850", "cpc_table"),
    "u850_centre": ("https://www.cpc.ncep.noaa.gov/data/indices/cpac850", "cpc_table"),
    "u850_east":  ("https://www.cpc.ncep.noaa.gov/data/indices/epac850", "cpc_table"),

    # ── ТОПЛИВО. Тёплый объём воды в верхних 300 метрах опережает Niño 3.4 на два-три сезона:
    # это единственный ИЗМЕРЯЕМЫЙ признак того, есть ли событию чем расти. До сих пор оценка
    # пика бралась только из аналогов и моделей.
    "wwv":        ("https://www.pmel.noaa.gov/tao/wwv/data/wwv.dat", "pmel"),
    "t300":       ("https://www.pmel.noaa.gov/tao/wwv/data/t300.dat", "pmel"),

    # ── СЛОИ АТМОСФЕРЫ (спутник, UAH v6.1). Тропосфера догревается через несколько месяцев
    # после океана, стратосфера, наоборот, стынет — видно, как волна идёт вверх по этажам.
    "uah_tlt":    ("https://www.nsstc.uah.edu/data/msu/v6.1/tlt/uahncdc_lt_6.1.txt", "uah"),
    "uah_tmt":    ("https://www.nsstc.uah.edu/data/msu/v6.1/tmt/uahncdc_mt_6.1.txt", "uah"),
    "uah_ttp":    ("https://www.nsstc.uah.edu/data/msu/v6.1/ttp/uahncdc_tp_6.1.txt", "uah"),
    "uah_tls":    ("https://www.nsstc.uah.edu/data/msu/v6.1/tls/uahncdc_ls_6.1.txt", "uah"),

    # ── ЦЕНЫ ПОИМЁННО. Индекс FAO — одно число на всю еду; Эль-Ниньо бьёт по товарам
    # поимённо (пальмовое масло, рис, какао). Pink Sheet Всемирного банка: месячные цены,
    # бесплатно и без регистрации, свежее индекса FAO на месяц.
    "wb_pink": ("https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/"
                "related/CMO-Historical-Data-Monthly.xlsx", "xlsx"),

    # ── ЭКСПЕРТИЗА 04.09. RONI — относительный ONI (аномалия Niño 3.4 минус средняя по тропикам):
    # с февраля 2026 NOAA классифицирует события по нему, на тёплом фоне он ниже ONI на десятые.
    # MEI v2 — многомерный индекс из пяти полей, независимая сверка нашей «сцепки». DMI —
    # Индийский океан (IOD), второй по важности индекс для Залива и муссона Индии. OMI —
    # индекс MJO от PSL (BoM закрыт для скриптов): западные всплески ветра кучкуются в фазах
    # 6–8. Теплосодержание океана NCEI — фон системы, квартально.
    "roni":     ("https://www.cpc.ncep.noaa.gov/data/indices/RONI.ascii.txt", "roni_txt"),
    "mei":      ("https://psl.noaa.gov/enso/mei/data/meiv2.data", "psl_monthly"),
    "dmi":      ("https://psl.noaa.gov/gcos_wgsp/Timeseries/Data/dmi.had.long.data", "psl_monthly"),
    # ROMI — OMI в реальном времени по OLR от CPC (omi.1x.txt у PSL отстаёт на месяцы: проверено
    # 04.09, конец июня). Колонки: год месяц день флаг PC1 PC2 амплитуда.
    "omi":      ("https://psl.noaa.gov/mjo/mjoindex/romi.cpcolr.1x.txt", "omi_txt"),
    "kuwait_era5": (kuwait_url, "json"),
}
# Теплосодержание NCEI лежит по кварталам в четырёх файлах на каждую глубину (файл без суффикса
# в том же каталоге — старый годовой ряд до 2010-го, проверено 04.09). Тянем все восемь.
_OHC = "https://www.ncei.noaa.gov/data/oceans/woa/DATA_ANALYSIS/3M_HEAT_CONTENT/DATA/basin/3month/"
OHC_QUARTERS = ("1-3", "4-6", "7-9", "10-12")
for _d in (700, 2000):
    for _q in OHC_QUARTERS:
        SOURCES[f"ohc_{_d}_{_q}"] = (f"{_OHC}h22-w0-{_d}m{_q}.dat", "ncei_ohc")

LABELS = {
    "t2_world": "Land+ocean, world (ERA5, 2 m)",
    "t2_nh": "Land+ocean, northern hemisphere (ERA5)",
    "t2_sh": "Land+ocean, southern hemisphere (ERA5)",
    "sst_world": "Ocean, 60°S–60°N (OISST)",
    "sst_nino34": "Niño 3.4, the El Niño focus (OISST)",
    "noaa_weekly": "NOAA weekly indices: Niño 1+2, 3, 3.4, 4",
    "oni": "ONI, NOAA official 3-month index",
    "psl_nino34_monthly": "Niño 3.4 monthly (ERSST v6, PSL)",
    "fao_fpi": "FAO Food Price Index and five groups (monthly)",
    "soi": "Southern Oscillation Index (Tahiti minus Darwin)",
    "olr": "Outgoing longwave radiation at the date line",
    "u850_west": "Trade wind at 850 hPa, western Pacific",
    "u850_centre": "Trade wind at 850 hPa, central Pacific",
    "u850_east": "Trade wind at 850 hPa, eastern Pacific",
    "wwv": "Warm water volume, equatorial Pacific (PMEL)",
    "t300": "Upper 300 m temperature, equatorial Pacific (PMEL)",
    "uah_tlt": "Satellite temperature, lower troposphere (UAH)",
    "uah_tmt": "Satellite temperature, mid troposphere (UAH)",
    "uah_ttp": "Satellite temperature, tropopause (UAH)",
    "uah_tls": "Satellite temperature, lower stratosphere (UAH)",
    "wb_pink": "World Bank Pink Sheet: monthly commodity prices",
    "roni": "RONI, relative ONI (NOAA CPC, official since Feb 2026)",
    "mei": "Multivariate ENSO Index v2 (NOAA PSL)",
    "dmi": "Dipole Mode Index, Indian Ocean (HadISST via PSL)",
    "omi": "ROMI, the real-time MJO index (NOAA PSL), daily",
    "kuwait_era5": "Kuwait, ERA5 daily at 29.37°N 47.98°E (Open-Meteo)",
}
for _d in (700, 2000):
    for _q in OHC_QUARTERS:
        LABELS[f"ohc_{_d}_{_q}"] = f"Ocean heat content 0–{_d} m, months {_q} (NCEI)"


def ext_of(kind):
    """Расширение файла по виду источника. Одно место на всех: кэшированный путь в watch.run
    строил его сам, и с приходом xlsx начал искать wb_pink.txt, которого не бывает."""
    return {"cr_json": ".json", "json": ".json", "fao_csv": ".csv", "xlsx": ".xlsx"}.get(kind, ".txt")


def fetch_all(timeout=40):
    """Тянет всё, что может; возвращает {name: (path, fresh: bool, error)}."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = RAW / stamp
    dest.mkdir(parents=True, exist_ok=True)
    LAST.mkdir(parents=True, exist_ok=True)
    out = {}
    for name, (url, kind) in SOURCES.items():
        ext = ext_of(kind)
        p = dest / (name + ext)
        try:
            if callable(url):
                url = url()
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                data = r.read()
            if kind in ("cr_json", "json"):
                json.loads(data.decode("utf-8"))      # проверка, что это JSON, а не страница ошибки
            elif len(data) < 500:
                raise ValueError("слишком короткий ответ: %r" % data[:80])
            p.write_bytes(data)
            shutil.copy(p, LAST / (name + ext))
            out[name] = (LAST / (name + ext), True, "")
        except Exception as e:                       # noqa: BLE001 - любая сетевая беда
            lp = LAST / (name + ext)
            out[name] = (lp if lp.exists() else None, False, str(e)[:160])
    # пустой каталог штампа, если всё упало, не нужен
    if not any(p.exists() for p in dest.iterdir()):
        dest.rmdir()
    return out, stamp


# ------------------------------------------------------------------ разбор
def read_cpc_table(path, table=-1):
    """Таблицы CPC: «YEAR JAN … DEC», в одном файле их две-три подряд.

    В SOI это АНОМАЛИЯ и СТАНДАРТИЗОВАННАЯ, в OLR и ветрах — ИСХОДНАЯ, АНОМАЛИЯ и
    СТАНДАРТИЗОВАННАЯ. Нам почти всегда нужна последняя (table=-1): в стандартизованном виде
    ряды сравнимы между собой, а «−2 сигмы» читается одинаково и для давления, и для ветра.
    Колонки ФИКСИРОВАННОЙ ширины, и это важно: пропуски пишутся как -999.9 и слипаются
    («-999.9-999.9»), поэтому разбивать по пробелам нельзя — только по позициям.
    """
    rows, tables = [], []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.strip().startswith("YEAR"):
            if rows:
                tables.append(rows)
            rows = []
            continue
        if len(line) < 16 or not line[:4].strip().isdigit():
            continue
        year = int(line[:4])
        for m in range(12):
            chunk = line[4 + m * 6:10 + m * 6].strip()
            if not chunk:
                continue
            try:
                v = float(chunk)
            except ValueError:
                continue
            if v <= -999:
                continue
            rows.append((f"{year}-{m + 1:02d}", v))
    if rows:
        tables.append(rows)
    if not tables:
        return {}
    return dict(tables[table])


def read_pmel(path):
    """PMEL: дата YYYYMM, объём, аномалия — в научной записи. Берём аномалию."""
    out = {}
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        f = line.split()
        if len(f) < 3 or not f[0].isdigit() or len(f[0]) != 6:
            continue
        try:
            out[f"{f[0][:4]}-{f[0][4:]}"] = float(f[2])
        except ValueError:
            continue
    return out


def read_uah(path):
    """UAH: «Year Mo Globe Land Ocean NH … Trpcs …». Берём глобальную и тропическую колонки.

    Тропики — то, что нас касается: сигнал Эль-Ниньо в свободной атмосфере живёт именно там,
    а в глобальном среднем он размазан."""
    globe, trop = {}, {}
    cols = None
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        f = line.split()
        if not f:
            continue
        if f[0] == "Year":
            cols = f
            continue
        if not f[0].isdigit() or len(f[0]) != 4 or len(f) < 12:
            continue
        key = f"{int(f[0])}-{int(f[1]):02d}"
        try:
            globe[key] = float(f[2])
            ti = cols.index("Trpcs") if cols and "Trpcs" in cols else 11
            trop[key] = float(f[ti])
        except (ValueError, IndexError):
            continue
    return {"globe": globe, "tropics": trop}


# Что именно берём из Pink Sheet: товары, по которым Эль-Ниньо бьёт по литературе и по
# прошлым событиям. Полный лист — сотня колонок, и большая часть к делу не относится.
PINK_PICK = {
    "Palm oil": "palm_oil", "Soybean oil": "soybean_oil", "Coconut oil": "coconut_oil",
    "Rice, Thai 5%": "rice", "Wheat, US HRW": "wheat", "Maize": "maize",
    "Sugar, world": "sugar", "Coffee, Arabica": "coffee_arabica", "Cocoa": "cocoa",
    "Fish meal": "fishmeal", "DAP": "fertilizer_dap", "Urea": "fertilizer_urea",
}


def read_pink(path):
    """Pink Sheet Всемирного банка: лист «Monthly Prices», шапка в две строки.

    Строка 5 — названия товаров, строка 6 — единицы, данные с седьмой; первая колонка вида
    «2026M08». Имена в шапке иногда меняются («Rice, Thai 5%» ↔ «Rice, Thai 5% »), поэтому
    сверяем по началу строки без регистра, а не по точному совпадению."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Monthly Prices" not in wb.sheetnames:
        return {}
    ws = wb["Monthly Prices"]
    rows = ws.iter_rows(values_only=True)
    head, unit = None, None
    out, months = {}, []
    for i, r in enumerate(rows, 1):
        if i == 5:
            head = [str(c).strip() if c else "" for c in r]
        elif i == 6:
            unit = [str(c).strip() if c else "" for c in r]
        elif i > 6 and r and isinstance(r[0], str) and "M" in r[0]:
            y, m = r[0].split("M")
            months.append(f"{y}-{int(m):02d}")
            for ci, name in enumerate(head or []):
                key = None
                for want, short in PINK_PICK.items():
                    if name.lower().startswith(want.lower()):
                        key = short
                        break
                if not key:
                    continue
                v = r[ci] if ci < len(r) else None
                rec = out.setdefault(key, {"name": name, "unit": (unit or [""] * (ci + 1))[ci], "series": {}})
                if isinstance(v, (int, float)):
                    rec["series"][months[-1]] = round(float(v), 3)
    return out


def read_cr_json(path):
    """Ряды climatereanalyzer: {год: массив 366}, климатология 1991-2020, дата последнего дня."""
    d = json.loads(Path(path).read_text(encoding="utf-8"))
    years, clim = {}, None
    for r in d:
        nm = r.get("name", "")
        arr = np.array([np.nan if v is None else v for v in r["data"]], float)
        if nm.isdigit():
            years[int(nm)] = arr
        elif nm == "1991-2020":
            clim = arr
    y_last = max(years)
    fin = np.where(np.isfinite(years[y_last]))[0]
    n = int(len(fin))
    last_idx = int(fin[-1]) if n else -1
    last_day = grid_index_to_date(y_last, last_idx) if n else None
    return {"years": years, "clim": clim, "last_year": y_last,
            "last_n": n, "last_idx": last_idx, "last_date": last_day}


def grid_index_to_date(year, idx):
    """Индекс в 366-дневной сетке ряда -> реальная дата. Сетка всегда содержит
    29 февраля (индекс 59); в невисокосный год эта ячейка пуста и её надо перешагнуть."""
    from datetime import timedelta
    import calendar
    if calendar.isleap(year) or idx < 59:
        return date(year, 1, 1) + timedelta(days=idx)
    return date(year, 1, 1) + timedelta(days=idx - 1)


def read_noaa_weekly(path):
    """Недельные SST и аномалии по четырём регионам Niño."""
    rows = []
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        m = re.match(r"\s*(\d{2}[A-Z]{3}\d{4})\s+(.*)", ln)
        if not m:
            continue
        dt = datetime.strptime(m.group(1), "%d%b%Y").date()
        nums = re.findall(r"-?\d+\.\d", m.group(2))
        if len(nums) < 8:
            continue
        v = list(map(float, nums[:8]))
        rows.append({"date": dt, "n12": v[0], "n12a": v[1], "n3": v[2], "n3a": v[3],
                     "n34": v[4], "n34a": v[5], "n4": v[6], "n4a": v[7]})
    return rows


def read_oni(path):
    """ONI из oni.ascii.txt: список (сезон, год, аномалия) по порядку, DJF..NDJ."""
    out = []
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        parts = ln.split()
        if len(parts) == 4 and re.fullmatch(r"[A-Z]{3}", parts[0]) and parts[1].isdigit():
            out.append((parts[0], int(parts[1]), float(parts[3])))
    return out


def read_roni(path):
    """RONI из RONI.ascii.txt: список (сезон, год, аномалия) — как read_oni, но без колонки TOTAL."""
    out = []
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        parts = ln.split()
        if len(parts) == 3 and re.fullmatch(r"[A-Z]{3}", parts[0]) and parts[1].isdigit():
            try:
                out.append((parts[0], int(parts[1]), float(parts[2])))
            except ValueError:
                continue
    return out


def read_omi(path):
    """OMI от PSL: «год месяц день PC1 PC2 амплитуда» по дням с 1991-го. {дата: (pc1, pc2, amp)}."""
    out = {}
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        f = ln.split()
        if len(f) < 6 or not f[0].isdigit():
            continue
        try:
            if len(f) >= 7:                      # ROMI: год месяц день флаг PC1 PC2 амплитуда
                out[f"{int(f[0])}-{int(f[1]):02d}-{int(f[2]):02d}"] = (float(f[4]), float(f[5]), float(f[6]))
            else:                                # OMI: год месяц день PC1 PC2 амплитуда
                out[f"{int(f[0])}-{int(f[1]):02d}-{int(f[2]):02d}"] = (float(f[3]), float(f[4]), float(f[5]))
        except ValueError:
            continue
    return out


def read_ohc(path):
    """Теплосодержание NCEI: «YEAR WO WOse NH NHse SH SHse», год — дробный (середина квартала),
    единицы 10²² Дж от среднего 1955–2006. Берём мировое (WO)."""
    out = []
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        f = ln.split()
        if len(f) < 3 or not re.fullmatch(r"\d{4}\.\d+", f[0]):
            continue
        try:
            out.append((float(f[0]), float(f[1]), float(f[2])))
        except ValueError:
            continue
    return out


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def read_psl_monthly(path):
    """Месячная аномалия Niño 3.4 (ERSST): {год: [12]}; -99.99 = нет данных."""
    out = {}
    for ln in io.open(path, encoding="utf-8", errors="replace"):
        parts = ln.split()
        if len(parts) == 13 and re.fullmatch(r"\d{4}", parts[0]):
            vals = [float(x) for x in parts[1:]]
            out[int(parts[0])] = [np.nan if v < -90 else v for v in vals]
    return out


if __name__ == "__main__":
    res, stamp = fetch_all()
    for k, (p, fresh, err) in res.items():
        print("%-20s %s %s" % (k, "свежее" if fresh else "СТАРОЕ", err))


def read_fao(path):
    """FAO Food Price Index: {"months": ["1990-01", ...], "index": [...], "groups": {"Meat": [...], ...}}.
    Строки без числа в индексе пропускаем; последняя строка файла — последний вышедший месяц."""
    import csv
    months, index, groups = [], [], {"Meat": [], "Dairy": [], "Cereals": [], "Oils": [], "Sugar": []}
    with io.open(path, encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.reader(f))
    head = next((r for r in rows if r and r[0].strip() == "Date"), None)
    if not head:
        raise ValueError("FAO CSV: строка заголовка Date не найдена — формат изменился")
    col = {name.strip(): i for i, name in enumerate(head) if name.strip()}
    for r in rows:
        if not r or not re.match(r"^\d{4}-\d{2}$", (r[0] or "").strip()):
            continue
        try:
            v = float(r[col["Food Price Index"]])
        except (ValueError, IndexError, KeyError):
            continue
        months.append(r[0].strip()); index.append(v)
        for g in groups:
            try:
                groups[g].append(float(r[col[g]]))
            except (ValueError, IndexError, KeyError):
                groups[g].append(None)
    return {"months": months, "index": index, "groups": groups}
